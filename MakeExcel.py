import os
import openpyxl

def RifleExcel(Info):
    # print("Weapon name:", Info[0])
    # print("Damage:", Info[1])
    # print("Multishot:", Info[2])
    # print("Fire Rate:", Info[3])
    # print("Critical Chance:", Info[4])
    # print("Critical Damage:", Info[5])
    # print("Status Chance:", Info[6])
    wb = openpyxl.Workbook()
    ws = wb.active
    # Tikai neuztraucieties, es šo visu nerakstīju ar roku :), es tam uztaisīju automatizāciju
    ws.column_dimensions["A"].width = 21
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["D"].width = 11
    ws.column_dimensions["I"].width = 21
    ws.column_dimensions["R"].width = 11
    ws.column_dimensions["W"].width = 21
    ws["A1"] = Info[0]
    ws["A3"] = Info[1]
    ws["A4"] = Info[2]
    ws["A5"] = Info[3]
    ws["A6"] = Info[4]
    ws["A7"] = Info[5]
    ws["A8"] = Info[6]
    ws["A2"] = "Unmodded"
    ws["A10"] = "Mods list"
    ws["A11"] = "Serration"
    ws["A12"] = "Amalgam Serration"
    ws["A13"] = "Galvanized Chamber"
    ws["A14"] = "Vigialnte Armaments"
    ws["A15"] = "Primed Shred"
    ws["A16"] = "Vile acceleration"
    ws["A17"] = "Critical Delay"
    ws["A18"] = "Point Strike"
    ws["A19"] = "Vital Sense"
    ws["A20"] = "Hammer Shot"
    ws["A21"] = "Galvanized Aptitude"
    ws["A22"] = "Convulsion"
    ws["A23"] = "High Voltage"
    ws["A24"] = "Infected Clip"
    ws["A25"] = "Malignant Force"
    ws["A26"] = "Primed Cryo Rounds"
    ws["A27"] = "Rime Rounds"
    ws["A28"] = "Hellfire"
    ws["A29"] = "Thermite Rounds"
    ws["A30"] = "Hunter Munitions"
    ws["B2"] = "Stats"
    ws["B3"] = "dmg"
    ws["B4"] = "ms"
    ws["B5"] = "fr"
    ws["B6"] = "cc"
    ws["B7"] = "cd"
    ws["B8"] = "sc"
    ws["B10"] = "Stats"
    ws["B11"] = "1.65 dmg"
    ws["B12"] = "1.55 dmg"
    ws["B13"] = "0.8 ms"
    ws["B14"] = "0.6 ms"
    ws["B15"] = "0.55 fr"
    ws["B16"] = "0.9 fr, -0.15 dmg"
    ws["B17"] = "2.0 cc, -0.2 fr"
    ws["B18"] = "1.5 cc"
    ws["B19"] = "1.2 cd"
    ws["B20"] = "0.6 cd, 0.8 sc"
    ws["B21"] = "0.8 sc"
    ws["B22"] = "0.9 elec"
    ws["B23"] = "0.6 elec, 0.6 sc"
    ws["B24"] = "0.9 toxin"
    ws["B25"] = "0.6 toxin, 0.6 sc"
    ws["B26"] = "1.65 cold"
    ws["B27"] = "0.6 cold, 0.6 sc"
    ws["B28"] = "0.9 heat"
    ws["B29"] = "0.6 heat, 0.6 sc"
    ws["D10"] = "per shot"
    ws["D11"] = "per second"
    ws["D22"] = "per shot"
    ws["D23"] = "per second"
    ws["D34"] = "per shot"
    ws["D35"] = "per second"
    ws["E2"] = "Modded"
    ws["E3"] = "=$A$3*(1+SUM(J3:J11))*(1+SUM(P3:P11))"
    ws["E4"] = "=$A$4*(1+SUM(K3:K11))"
    ws["E5"] = "=$A$5*(1+SUM(L3:L11))"
    ws["E6"] = "=$A$6*(1+SUM(M3:M11))"
    ws["E7"] = "=$A$7*(1+SUM(N3:N11))"
    ws["E8"] = "=$A$8*(1+SUM(O3:O11))"
    ws["E9"] = "avg dmg"
    ws["E10"] = "=E3*E4*(1+E6*(E7-1))"
    ws["E11"] = "=E10*E5"
    ws["E14"] = "Modded"
    ws["E15"] = "=$A$3*(1+SUM(J15:J23))*(1+SUM(P15:P23))"
    ws["E16"] = "=$A$4*(1+SUM(K15:K23))"
    ws["E17"] = "=$A$5*(1+SUM(L15:L23))"
    ws["E18"] = "=$A$6*(1+SUM(M15:M23))"
    ws["E19"] = "=$A$7*(1+SUM(N15:N23))"
    ws["E20"] = "=$A$8*(1+SUM(O15:O23))"
    ws["E21"] = "avg dmg"
    ws["E22"] = "=E15*E16*(1+E18*(E19-1))"
    ws["E23"] = "=E22*E17"
    ws["E26"] = "Modded"
    ws["E27"] = "=$A$3*(1+SUM(J27:J35))*(1+SUM(P27:P35))"
    ws["E28"] = "=$A$4*(1+SUM(K27:K35))"
    ws["E29"] = "=$A$5*(1+SUM(L27:L35))"
    ws["E30"] = "=$A$6*(1+SUM(M27:M35))"
    ws["E31"] = "=$A$7*(1+SUM(N27:N35))"
    ws["E32"] = "=$A$8*(1+SUM(O27:O35))"
    ws["E33"] = "avg dmg"
    ws["E34"] = "=E27*E28*(1+E30*(E31-1))"
    ws["E35"] = "=E34*E29"
    ws["F2"] = "Stats"
    ws["F3"] = "dmg"
    ws["F4"] = "ms"
    ws["F5"] = "fr"
    ws["F6"] = "cc"
    ws["F7"] = "cd"
    ws["F8"] = "sc"
    ws["F9"] = "avg sc"
    ws["F10"] = "=E4*E8"
    ws["F11"] = "=F10*E5"
    ws["F14"] = "Stats"
    ws["F15"] = "dmg"
    ws["F16"] = "ms"
    ws["F17"] = "fr"
    ws["F18"] = "cc"
    ws["F19"] = "cd"
    ws["F20"] = "sc"
    ws["F21"] = "avg sc"
    ws["F22"] = "=E16*E20"
    ws["F23"] = "=F22*E17"
    ws["F26"] = "Stats"
    ws["F27"] = "dmg"
    ws["F28"] = "ms"
    ws["F29"] = "fr"
    ws["F30"] = "cc"
    ws["F31"] = "cd"
    ws["F32"] = "sc"
    ws["F33"] = "avg sc"
    ws["F34"] = "=E28*E32"
    ws["F35"] = "=F34*E29"
    ws["G9"] = "avg sc*"
    ws["G10"] = "=IF(E6<1, E6*0.3*E4, 0.3*E4)"
    ws["G11"] = "=G10*E5"
    ws["G21"] = "avg sc*"
    ws["G22"] = "=$A$3*(1+SUM(J15:J23))*P22/E15*E20*E16"
    ws["G23"] = "=G22*E17"
    ws["G33"] = "avg sc*"
    ws["G34"] = "=$A$3*(1+SUM(J27:J35))*P34/E27*E32*E28"
    ws["G35"] = "=G34*E29"
    ws["I1"] = "Viral + Hunter Munitions"
    ws["I2"] = "MODS"
    ws["I3"] = "Serration"
    ws["I4"] = "Galvanized Chamber"
    ws["I5"] = "Primed Shred"
    ws["I6"] = "Critical Delay"
    ws["I7"] = "Vital Sense"
    ws["I8"] = "cold"
    ws["I9"] = "Malignant Force"
    ws["I10"] = "Hunter Munitions*"
    ws["I11"] = "Other Buffs"
    ws["I13"] = "Corrosive + Cold (Maximum raw damage)"
    ws["I14"] = "MODS"
    ws["I15"] = "Serration"
    ws["I16"] = "Galvanized Chamber"
    ws["I17"] = "Primed Shred"
    ws["I18"] = "Critical Delay"
    ws["I19"] = "Vital Sense"
    ws["I20"] = "Convulsion"
    ws["I21"] = "Malignant Force"
    ws["I22"] = "Primed Cryo Rounds*"
    ws["I23"] = "Other Buffs"
    ws["I25"] = "Corrosive + Cold (Status focused)"
    ws["I26"] = "MODS"
    ws["I27"] = "Serration"
    ws["I28"] = "Galvanized Chamber"
    ws["I29"] = "Primed Shred"
    ws["I30"] = "Critical Delay"
    ws["I31"] = "Hammer Shot"
    ws["I32"] = "High Voltage"
    ws["I33"] = "Infected Clip"
    ws["I34"] = "Primed Cryo Rounds*"
    ws["I35"] = "Other Buffs"
    ws["J2"] = "dmg"
    ws["J3"] = 1.65
    ws["J14"] = "dmg"
    ws["J15"] = 1.65
    ws["J26"] = "dmg"
    ws["J27"] = 1.65
    ws["K2"] = "ms"
    ws["K4"] = 0.8
    ws["K14"] = "ms"
    ws["K16"] = 0.8
    ws["K26"] = "ms"
    ws["K28"] = 0.8
    ws["L2"] = "fire rate"
    ws["L5"] = 0.55
    ws["L6"] = -0.2
    ws["L14"] = "fire rate"
    ws["L17"] = 0.55
    ws["L18"] = -0.2
    ws["L26"] = "fire rate"
    ws["L29"] = 0.55
    ws["L30"] = -0.2
    ws["M2"] = "cc"
    ws["M6"] = 2
    ws["M14"] = "cc"
    ws["M18"] = 2
    ws["M26"] = "cc"
    ws["M30"] = 2
    ws["N2"] = "cd"
    ws["N7"] = 1.2
    ws["N14"] = "cd"
    ws["N19"] = 1.2
    ws["N26"] = "cd"
    ws["N31"] = 0.6
    ws["O2"] = "sc"
    ws["O8"] = 0.6
    ws["O9"] = 0.6
    ws["O14"] = "sc"
    ws["O26"] = "sc"
    ws["O31"] = 0.8
    ws["O32"] = 0.6
    ws["O33"] = 0.6
    ws["P2"] = "element"
    ws["P8"] = 0.6
    ws["P9"] = 0.6
    ws["P14"] = "element"
    ws["P20"] = 0.9
    ws["P21"] = 0.9
    ws["P22"] = 1.65
    ws["P26"] = "element"
    ws["P32"] = 0.6
    ws["P33"] = 0.6
    ws["P34"] = 1.65
    ws["R10"] = "per shot"
    ws["R11"] = "per second"
    ws["R22"] = "per shot"
    ws["R23"] = "per second"
    ws["S2"] = "modded"
    ws["S3"] = "=$A$3*(1+SUM(X3:X11))*(1+SUM(AD3:AD11))"
    ws["S4"] = "=$A$4*(1+SUM(Y3:Y11))"
    ws["S5"] = "=$A$5*(1+SUM(Z3:Z11))"
    ws["S6"] = "=$A$6*(1+SUM(AA3:AA11))"
    ws["S7"] = "=$A$7*(1+SUM(AB3:AB11))"
    ws["S8"] = "=$A$8*(1+SUM(AC3:AC11))"
    ws["S9"] = "avg dmg"
    ws["S10"] = "=S3*S4*(1+S6*(S7-1))"
    ws["S11"] = "=S10*S5"
    ws["S14"] = "modded"
    ws["S15"] = "=$A$3*(1+SUM(X15:X23))*(1+SUM(AD15:AD23))"
    ws["S16"] = "=$A$4*(1+SUM(Y15:Y23))"
    ws["S17"] = "=$A$5*(1+SUM(Z15:Z23))"
    ws["S18"] = "=$A$6*(1+SUM(AA15:AA23))"
    ws["S19"] = "=$A$7*(1+SUM(AB15:AB23))"
    ws["S20"] = "=$A$8*(1+SUM(AC15:AC23))"
    ws["S21"] = "avg dmg"
    ws["S22"] = "=S15*S16*(1+S18*(S19-1))"
    ws["S23"] = "=S22*S17"
    ws["T2"] = "Stats"
    ws["T3"] = "dmg"
    ws["T4"] = "ms"
    ws["T5"] = "fr"
    ws["T6"] = "cc"
    ws["T7"] = "cd"
    ws["T8"] = "sc"
    ws["T9"] = "avg sc"
    ws["T10"] = "=S4*S8"
    ws["T11"] = "=T10*S5"
    ws["T14"] = "Stats"
    ws["T15"] = "dmg"
    ws["T16"] = "ms"
    ws["T17"] = "fr"
    ws["T18"] = "cc"
    ws["T19"] = "cd"
    ws["T20"] = "sc"
    ws["T21"] = "avg sc"
    ws["T22"] = "=S16*S20"
    ws["T23"] = "=T22*S17"
    ws["U9"] = "avg sc*"
    ws["U21"] = "avg sc*"
    ws["W1"] = "Custom Build 1"
    ws["W2"] = "MODS"
    ws["W11"] = "Other buffs"
    ws["W13"] = "Custom Build 2"
    ws["W14"] = "MODS"
    ws["W23"] = "Other buffs"
    ws["X2"] = "dmg"
    ws["X14"] = "dmg"
    ws["Y2"] = "ms"
    ws["Y14"] = "ms"
    ws["Z2"] = "fire rate"
    ws["Z14"] = "fire rate"
    ws["AA2"] = "cc"
    ws["AA14"] = "cc"
    ws["AB2"] = "cd"
    ws["AB14"] = "cd"
    ws["AC2"] = "sc"
    ws["AC14"] = "sc"
    ws["AD2"] = "element"
    ws["AD14"] = "element"

    if(os.path.exists(Info[0] + " AutoExcel.xlsx")):
        print("File already exists!")
    else:
        wb.save(Info[0] + " AutoExcel.xlsx")

def ShotgunExcel():
    pass

def PistolExcel():
    pass

def MeleeExcel():
    pass